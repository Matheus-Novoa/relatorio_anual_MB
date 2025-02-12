import pandas as pd
import os
from difflib import get_close_matches
from openpyxl import load_workbook, styles
import warnings
warnings.filterwarnings('ignore')



# Read the Excel file, specifically the "Serviços" sheet
df = pd.read_excel('Serviços maple bear zona norte.xls', sheet_name='Serviços', header=None)

matriculas_df = pd.read_excel(
    'Matriculas.xlsx',
    sheet_name="Alunos Matriculados (Paulo)",
    usecols=['Aluno', 'Responsável Financeiro', 'CPF', 'Endereço', 'Bairro']
)

# Rename columns
df.columns = ['Data', 'Nota', 'Cliente', 'Valor Contábil'] if len(df.columns) >= 4 else df.columns

# Remove rows with empty dates or where 'Data' is in the Date column
df = df[df['Data'].notna()]
df = df[df['Data'].astype(str).str.lower() != 'data']

# Get unique client names
unique_clients = df['Cliente'].unique()

# Create output directory if it doesn't exist
output_dir = 'extracted_tables'
os.makedirs(output_dir, exist_ok=True)

# Extract tables for each client
for client in unique_clients:
    if pd.notna(client):  # Skip if client name is NaN
        # Get rows for current client
        client_df = df[df['Cliente'] == client]

        # Convert 'Data' column to format dd/mm/yyyy
        client_df['Data'] = pd.to_datetime(client_df['Data'], errors='coerce').dt.strftime('%d/%m/%Y')
        
        nomes_pais = matriculas_df['Responsável Financeiro'].dropna().str.lower().tolist()
        nome_correspondente = get_close_matches(client.lower(), nomes_pais, n=1, cutoff=0.4)[0]
        
        client_df['CPF'] = matriculas_df[
            matriculas_df['Responsável Financeiro'].str.lower() == nome_correspondente
        ]['CPF'].values[0]

        dados_responsavel = matriculas_df[
            matriculas_df['Responsável Financeiro'].str.lower() == nome_correspondente
        ][['Aluno', 'Endereço', 'Bairro']].iloc[0:1]
        
        if not client_df.empty:
            # Create safe filename by removing invalid characters
            safe_filename = "".join(x for x in str(client) if x.isalnum() or x in (' ', '-', '_'))
            
            # Reordenar as colunas antes de salvar
            client_df = client_df[['Data', 'Nota', 'Cliente', 'CPF', 'Valor Contábil']]
            
            # Count occurrences of each date
            date_counts = client_df['Data'].value_counts()
            has_duplicates = (date_counts > 1).any()

            outputs = []
            
            if has_duplicates:
                duplicate_rows = client_df[client_df.duplicated(subset=['Data'], keep='first')]
                complementary_rows = client_df[~client_df.index.isin(duplicate_rows.index)]
                
                if not duplicate_rows.empty:
                    outputs.append(os.path.join(output_dir, f'{safe_filename}_filho1.xlsx'))
                    with pd.ExcelWriter(outputs[0], engine='openpyxl') as writer:
                        duplicate_rows.to_excel(writer, sheet_name='notas', index=False)
                        dados_responsavel.to_excel(writer, sheet_name='dados', index=False)
                
                if not complementary_rows.empty:
                    outputs.append(os.path.join(output_dir, f'{safe_filename}_filho2.xlsx'))
                    with pd.ExcelWriter(outputs[1], engine='openpyxl') as writer:
                        complementary_rows.to_excel(writer, sheet_name='notas', index=False)
                        dados_responsavel.to_excel(writer, sheet_name='dados', index=False)
            else:
                # Save single file if no duplicates
                outputs.append(os.path.join(output_dir, f'{safe_filename}.xlsx'))
                with pd.ExcelWriter(outputs[0], engine='openpyxl') as writer:
                    client_df.to_excel(writer, sheet_name='notas', index=False)
                    dados_responsavel.to_excel(writer, sheet_name='dados', index=False)

        for file in outputs:
            workbook = load_workbook(file)
            sheet = workbook['notas']
            
            client_df_from_excel = pd.read_excel(file, sheet_name='notas')
            total = client_df_from_excel['Valor Contábil'].sum()
            
            # Adicionar e formatar o total
            total_row = len(client_df_from_excel) + 2
            sheet.cell(row=total_row, column=1, value='Total')
            total_cell = sheet.cell(row=total_row, column=5, value=total)
            
            # Salvar as alterações
            workbook.save(file)
            print(f'Updated formatting in file: {file}')

